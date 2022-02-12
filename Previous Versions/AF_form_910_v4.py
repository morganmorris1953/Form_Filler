##in v3, i created methods to call in the GUI program

#TODO
#create variables (arrays) to search by name and by rank
#make the while loop more robust
# change input of names/rows to be used

from ctypes import pythonapi
from sys import setswitchinterval
import pyautogui, time
import os
import tkinter as tk
from tkinter import ttk
import openpyxl
import datetime
from datetime import datetime
# from GUI import rankOrName
# print(rankOrName)
# import GUI
pyautogui.FAILSAFE = True

# NORM_FONT = ("Helvetica", 10)
def Write_AF_form_910(rankOrName, nameList, rankList):
    delayTime = 3
    interval_time = 0.0
    pdfName = 'main_af-form-910-enlisted-performance-report-ab-thru-tsgt.pdf'
    pdfCommand = 'start ' + pdfName
    path = r"C:\Users\skype\OneDrive\Documents\fillOutPDF_folder\ALPHA_ROSTER_FIELDS.xlsm"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active


#######TODO#############
#determine if it's rank or name
#make a for loop to go through the array
# make another for loop for the entire program with the length of the array
# search for the appropriate characteristics to create the forms
# thingToSearch = rankOrName + "List"




    i = 2
    while (sheet_obj.cell(row = i, column = 1)).value != None:
        
        os.system(pdfCommand)
        name = (sheet_obj.cell(row = i, column = 1)).value
        ssn = sheet_obj.cell(row = i, column = 2)
        grade = (sheet_obj.cell(row = i, column = 3)).value

        ranks = {
            'AB': 1,
            'AMN': 2,
            'A1C': 3,
            'SRA': 4,
            'SGT': 6,
            'TSG': 8
        }
    
        # get() method of dictionary data type returns
        # value of passed argument if it is present
        # in dictionary otherwise second argument will
        # be assigned as default value of passed argument
        grade_key_presses = ranks.get(grade, 0)






        dafsc = (sheet_obj.cell(row = i, column = 11)).value
        command = "24 Special Tactics Squadron"
        pas = (sheet_obj.cell(row = i, column = 5)).value
        srid = '9999'
        report_start = ((sheet_obj.cell(row = i, column = 38)).value).date()
        report_end = ((sheet_obj.cell(row = i, column = 32)).value).date()
        supervisor = sheet_obj.cell(row=i, column=28)
        supervisor = supervisor.value
        ssn = ssn.value   #gets value from the ssn field
        ssn = ssn.replace("-", "")  #removes dashes from the ssn   #<class 'openpyxl.cell.cell.Cell'>
        date1 = report_start
        report_start = str(report_start)   #<class 'str'>
        new_report_start = datetime.strptime(report_start,'%Y-%m-%d').strftime('%d-%b-%Y')
        date2 = report_end
        report_end = str(report_end)
        new_report_end = datetime.strptime(report_end,'%Y-%m-%d').strftime('%d-%b-%Y')
        
        days_non_rated = 0
        days_supervised = (date2 - date1).days
        duty_title = (sheet_obj.cell(row = i, column = 8)).value
        
    
        
        #find supervisor information:
        # for j in range(2, 4):
        j = 2
        while (sheet_obj.cell(row = i, column = 1)).value != "":
            print((sheet_obj.cell(row = i, column = 1)).value)
            supervisor_name = ((sheet_obj.cell(row=j, column=1)).value)
            shortened_supervisor_name = supervisor_name.replace(',', '').rsplit(' ', 1)[0]
            if shortened_supervisor_name != supervisor:
                pass
            else:
                supervisor_rank = (sheet_obj.cell(row=j, column=3)).value
                supervisor_ssn = (sheet_obj.cell(row=j, column=2)).value
                supervisor_last_four = repr(supervisor_ssn)[-5:-1]    #the last position is a quote mark because supervisor_last_four is a string
                supervisor_information = ("%s, %s, USAF\n24 Special Tactics Squadron, AFSOC, Pope AAF, NC" % (supervisor_name, supervisor_rank))
                supervisor_duty_title = (sheet_obj.cell(row = j, column = 8).value)
                break  
            j += 1
        time.sleep(delayTime*3)


        #### write everything to the PDF ######
        pyautogui.write(f'%s\t' % (name), interval=interval_time)
        pyautogui.write('%s\t' % (ssn), interval=interval_time)
        pyautogui.press('down', presses=grade_key_presses, interval=interval_time)
        pyautogui.press('tab', interval=interval_time)
        pyautogui.write('%s\t' % (dafsc), interval=interval_time)
        pyautogui.write('%s\t' % (command), interval=interval_time)
        pyautogui.write('%s\t' % (pas), interval=interval_time)
        pyautogui.write('%s\t' % (srid), interval=interval_time)
        pyautogui.write('%s\t' % (new_report_start), interval=interval_time)
        pyautogui.write('%s\t' % (new_report_end), interval=interval_time)
        pyautogui.sleep(2)
        pyautogui.press('enter', interval=interval_time)
        pyautogui.write('%s\t' % (days_non_rated), interval=interval_time)
        pyautogui.write('%s\t' % (days_supervised), interval=interval_time)

        pyautogui.press('down', interval=interval_time)       #sets reason for report to 'annual'.  
        pyautogui.press('tab', interval=interval_time)
        pyautogui.write('%s' % (duty_title), interval=interval_time)
        pyautogui.press('tab', presses=4, interval=interval_time)
        pyautogui.press('space, interval=interval_time')         
        pyautogui.press('tab', presses=6, interval=interval_time)
        pyautogui.press('space', interval=interval_time)         
        pyautogui.press('tab', presses=6, interval=interval_time)
        pyautogui.press('space', interval=interval_time) 
        pyautogui.press('tab', presses=6, interval=interval_time)
        pyautogui.press('space', interval=interval_time) 
        pyautogui.press('tab', presses=3, interval=interval_time)

        pyautogui.write('%s\t' % (supervisor_information), interval=interval_time)
        pyautogui.write('%s\t\t' % (supervisor_duty_title), interval=interval_time)
        pyautogui.write('%s' % (supervisor_last_four), interval=interval_time)
    


        #save the document
        pyautogui.hotkey('ctrl', 's')
        pyautogui.sleep(delayTime)
        pyautogui.press('enter')
        pyautogui.sleep(delayTime)
        pyautogui.write(name + ' EPR')
        pyautogui.press('enter')
        pyautogui.hotkey('ctrl', 'f4')

        i += 1




# def OperationCompleteMessage(msg):
#     popup = tk.Tk()
#     popup.wm_title("!")
#     label = ttk.Label(popup, text=msg, font=NORM_FONT)
#     label.pack(side="top", fill="x", pady=100, padx=500)
#     B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
#     B1.pack()
#     popup.mainloop()

# OperationCompleteMessage("Operation Complete!")





## mess around with the while statements.  everything appears to be working fine, but after the last EPR, the program tries to pull up a blank form and fill it with blank info