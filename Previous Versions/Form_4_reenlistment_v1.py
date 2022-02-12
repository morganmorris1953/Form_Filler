#TODO
#create the GUI to select by name for reenlistment forms.





from ctypes import pythonapi
from sys import setswitchinterval
import pyautogui, time
import os
import tkinter as tk
from tkinter import ttk
# import csv
import openpyxl
import datetime
from datetime import datetime
# from datetime import date
pyautogui.FAILSAFE = True

NORM_FONT = ("Helvetica", 10)
delayTime = 3
interval_delay = 0.02

pdfName = 'Reenlistment_form_4.pdf'
pdfCommand = 'start ' + pdfName


path = r"C:\Users\skype\OneDrive\Documents\fillOutPDF_folder\ALPHA_ROSTER_FIELDS.xlsm"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

name = ''

searched_name = pyautogui.prompt('Who needs to reenlist?')
print(searched_name)

i = 2
while (sheet_obj.cell(row = i, column = 1)).value != "":
    name = (sheet_obj.cell(row = i, column = 1)).value
    if name == searched_name:
        break
    else:
        pass
    i += 1


# pyautogui.alert('This displays some text with an OK button.')  

os.system(pdfCommand)

ssn = sheet_obj.cell(row = i, column = 2).value
home_address = sheet_obj.cell(row = i, column = 24).value
home_city = sheet_obj.cell(row = i, column = 25).value
home_state = sheet_obj.cell(row = i, column = 26).value
home_zip = sheet_obj.cell(row = i, column = 27).value
home_of_record = (f"%s %s, %s %s" % (home_address, home_city, home_state, home_zip))
reenlistment_location = 'Camp Lemonnier, Djibouti'
reenlistment_date = '30 Nov 2021'
dob = str((sheet_obj.cell(row = i, column = 23).value).date())
datetime.strptime(dob,'%Y-%m-%d').strftime('%Y%m%d')
service_branch = "United States Air Force"
grade = (sheet_obj.cell(row = i, column = 3)).value

ranks = {
    'AB': 'E-1',
    'AMN': 'E-2',
    'A1C': 'E-3',
    'SRA': 'E-4',
    'SGT': 'E-5',
    'TSG': 'E-6',
    'MSG': 'E-7',
    'SMSG': 'E-8',
    'CMSG': 'E-9'
}
for rank in ranks:
    pay_grade = ranks[grade]


 # #### write everything to the PDF ######
pyautogui.sleep(delayTime*3)
pyautogui.press('tab', presses=2)
pyautogui.typewrite(f'%s\t' % (name))
pyautogui.write('%s\t' % (ssn), interval=interval_delay)
pyautogui.press('tab', interval=interval_delay)
pyautogui.write('%s\t' % (home_of_record), interval=interval_delay*2)
pyautogui.write('%s\t' % (reenlistment_location), interval=interval_delay)
pyautogui.press('down', presses=3)       #sets reason to 'reenlistment'.  
pyautogui.press('tab')
pyautogui.write('%s\t' % (reenlistment_date), interval=interval_delay)
pyautogui.write('%s' % (dob))
pyautogui.press('tab', presses=7, interval=interval_delay)
pyautogui.write('%s' % (service_branch), interval=interval_delay)
pyautogui.press('tab', presses=4, interval=interval_delay)
pyautogui.write('%s\t' % (pay_grade), interval=interval_delay)



#save the document
pyautogui.hotkey('ctrl', 's')
pyautogui.sleep(delayTime)
pyautogui.press('enter')
pyautogui.sleep(delayTime)
pyautogui.write(name + ' reenlistment')
pyautogui.press('enter')
# pyautogui.hotkey('ctrl', 'f4')




 
    # get() method of dictionary data type returns
    # value of passed argument if it is present
    # in dictionary otherwise second argument will
    # be assigned as default value of passed argument
    # grade_key_presses = ranks.get(grade, 0)






    # dafsc = (sheet_obj.cell(row = i, column = 11)).value
    # command = "24 Special Tactics Squadron"
    # pas = (sheet_obj.cell(row = i, column = 5)).value
    # srid = '9999'
    # report_start = ((sheet_obj.cell(row = i, column = 38)).value).date()
    # report_end = ((sheet_obj.cell(row = i, column = 32)).value).date()
    # supervisor = sheet_obj.cell(row=i, column=28)
    # supervisor = supervisor.value
    # ssn = ssn.value   #gets value from the ssn field
    # ssn = ssn.replace("-", "")  #removes dashes from the ssn   #<class 'openpyxl.cell.cell.Cell'>
    # date1 = report_start
    # report_start = str(report_start)   #<class 'str'>
    # new_report_start = datetime.strptime(report_start,'%Y-%m-%d').strftime('%d-%b-%Y')
    # date2 = report_end
    # report_end = str(report_end)
    # new_report_end = datetime.strptime(report_end,'%Y-%m-%d').strftime('%d-%b-%Y')
    
    # days_non_rated = 0
    # days_supervised = (date2 - date1).days
    # duty_title = (sheet_obj.cell(row = i, column = 8)).value
    
   
    
    # #find supervisor information:
    # # for j in range(2, 4):
    # j = 2
    # while (sheet_obj.cell(row = i, column = 1)).value != "":
    #     print((sheet_obj.cell(row = i, column = 1)).value)
    #     supervisor_name = ((sheet_obj.cell(row=j, column=1)).value)
    #     shortened_supervisor_name = supervisor_name.replace(',', '').rsplit(' ', 1)[0]
    #     if shortened_supervisor_name != supervisor:
    #         pass
    #     else:
    #         supervisor_rank = (sheet_obj.cell(row=j, column=3)).value
    #         supervisor_ssn = (sheet_obj.cell(row=j, column=2)).value
    #         supervisor_last_four = repr(supervisor_ssn)[-5:-1]    #the last position is a quote mark because supervisor_last_four is a string
    #         supervisor_information = ("%s, %s, USAF\n24 Special Tactics Squadron, AFSOC, Pope AAF, NC" % (supervisor_name, supervisor_rank))
    #         supervisor_duty_title = (sheet_obj.cell(row = j, column = 8).value)
    #         break  
    #     j += 1
    # time.sleep(delayTime)


    # #### write everything to the PDF ######
    # pyautogui.write(f'%s\t' % (name))
    # pyautogui.write('%s\t' % (ssn))
    # pyautogui.press('down', presses=grade_key_presses)
    # pyautogui.press('tab')
    # pyautogui.write('%s\t' % (dafsc))
    # pyautogui.write('%s\t' % (command))
    # pyautogui.write('%s\t' % (pas))
    # pyautogui.write('%s\t' % (srid))
    # pyautogui.write('%s\t' % (new_report_start))
    # pyautogui.write('%s\t' % (new_report_end))
    # pyautogui.sleep(2)
    # pyautogui.press('enter')
    # pyautogui.write('%s\t' % (days_non_rated))
    # pyautogui.write('%s\t' % (days_supervised))

    # pyautogui.press('down')       #sets reason for report to 'annual'.  
    # pyautogui.press('tab')
    # pyautogui.write('%s' % (duty_title))
    # pyautogui.press('tab', presses=4)
    # pyautogui.press('space')         
    # pyautogui.press('tab', presses=6)
    # pyautogui.press('space')         
    # pyautogui.press('tab', presses=6)
    # pyautogui.press('space') 
    # pyautogui.press('tab', presses=6)
    # pyautogui.press('space') 
    # pyautogui.press('tab', presses=3)

    # pyautogui.write('%s\t' % (supervisor_information))
    # pyautogui.write('%s\t\t' % (supervisor_duty_title))
    # pyautogui.write('%s' % (supervisor_last_four))
   


    # #save the document
    # pyautogui.hotkey('ctrl', 's')
    # pyautogui.sleep(4)
    # pyautogui.press('enter')
    # pyautogui.sleep(2)
    # pyautogui.write(name + ' EPR')
    # pyautogui.press('enter')
    # pyautogui.hotkey('ctrl', 'f4')

    # i += 1




# def OperationCompleteMessage(msg):
#     popup = tk.Tk()
#     popup.wm_title("!")
#     label = ttk.Label(popup, text=msg, font=NORM_FONT)
#     label.pack(side="top", fill="x", pady=100, padx=500)
#     B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
#     B1.pack()
#     popup.mainloop()

# OperationCompleteMessage("Operation Complete!")





## mess around with the while statements.  everything appears to be working fine, but after the last EPR, the program tries to pull up a blank form and fill it with blank info