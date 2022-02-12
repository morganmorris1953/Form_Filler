#! python3
import tkinter
from tkinter import OptionMenu, Tk, mainloop, TOP, ttk
from tkinter.ttk import Button
import openpyxl
import os

path = r"C:\Users\skype\OneDrive\Documents\fillOutPDF_folder\ALPHA_ROSTER_FIELDS.xlsm"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active



root = tkinter.Tk()
root.geometry('450x250')
root.title("Automated Form Filler")
root.config(bg='#F2B90C')


# 1. left frame
leftSideFrameVariable = tkinter.Frame(root)
leftSideFrameVariable.grid(row = 0, column = 0)
# leftSideFrameVariable.config(bg='#F2B90C')

buttonVariable = tkinter.IntVar()
optionVar = tkinter.StringVar()
forms = ('Form 910 (AB - TSgt EPR)', 'Form 911 (MSgt - SMSgt EPR)', 'Form 4 (Reenlistment)')


def display_selected(optionMenuValue):
    optionMenuValue = optionVar.get()
    print(optionMenuValue)
   

# def determine_program_to_run(optionMenuValue):
#     switcher = {
#         'Form 910 (AB - TSgt EPR)': 'AF_form_910_v2.py',
#         'Form 911 (MSgt - SMSgt EPR)': 'AF_form_911_v1.py',
#         'Form 4 (Reenlistment)': 'Form_4_reenlistment.py'
#         }
#     return switcher.get(optionMenuValue, "nothing")

formOptionMenu = ttk.OptionMenu(
    leftSideFrameVariable, 
    optionVar, 
    'Choose form', 
    *forms,
    command = display_selected) 
buttonLabel = tkinter.Label(
    leftSideFrameVariable,
    text = 'Choose:')
button1 = tkinter.Radiobutton(
    leftSideFrameVariable,
    text = 'By Rank',
    variable = buttonVariable,
    value = 1,
    justify='left',
    command=lambda: getScript(button1))
button2 = tkinter.Radiobutton(
    leftSideFrameVariable,
    text = 'By Name',
    variable = buttonVariable,
    value = 2,
    justify='left',
    command=lambda: getScript(button2))
runButton = tkinter.Button(
    leftSideFrameVariable,
    text = 'Run Program',
    command = lambda: runProgram(processingLabel))
processingLabel = tkinter.Label(
        leftSideFrameVariable,
        text = '')

#Left side button placement
formOptionMenu.grid(row = 0, column = 0, padx = 40, pady = 10)
button1.grid(row = 3, column = 0)
button2.grid(row = 4, column = 0)
runButton.grid(row = 5, column = 0, padx = 100, pady = 10)
processingLabel.grid(row = 6, column = 0)



# Run the program after choosing an option button
def runProgram(processingLabel):  #I think this should import the optionMenuValue and i can use it to start the correct program
    # self.optionMenuValue = optionMenuValue
    if buttonVariable.get() == 1:
        textDisplayed = 'Searching by rank'
    elif buttonVariable.get() == 2:
        textDisplayed = 'Search by name'        
        selectedItems = []
        for i in listBoxVariable.curselection():
            selectedItem = listBoxVariable.get(i)
            selectedItems.append(selectedItem)
        print(selectedItems)
   
    else:
        textDisplayed = 'You must choose either by rank or by name'
    processingLabel.configure(text = textDisplayed)
 

# 2. right frame
rightSideFrameVariable = tkinter.Frame(root)
rightSideFrameVariable.grid(row = 0, column = 1)


scrollbarVariable = tkinter.Scrollbar(
    rightSideFrameVariable,
    orient = tkinter.VERTICAL)
listBoxVariable = tkinter.Listbox(
    rightSideFrameVariable,
    width = 25,
    yscrollcommand = scrollbarVariable.set,
    selectmode = tkinter.EXTENDED)
scrollbarVariable.config(command = listBoxVariable.yview)
scrollbarVariable.pack(side = tkinter.RIGHT, fill = tkinter.Y)
listBoxVariable.pack()

# This method places either the ranks or the names into the field once either radio button is pressed.  Called as an argument by the radio buttons
def getScript(buttonPressed):
    listBoxVariable.delete(0, tkinter.END)
    if buttonPressed == button1:
        rankList = ['AB', 'A1C', 'SrA', 'SSgt', 'TSgt', 'MSgt', 'SMSgt']
        for j in range(len(rankList)):
            listBoxVariable.insert(tkinter.END, rankList[j])
    
    elif buttonPressed == button2:
        nameList = [] 
        i = 2
        while str((sheet_obj.cell(row = i, column = 1)).value) != '':
            currentName = str((sheet_obj.cell(row = i, column = 1)).value)
            if currentName == 'None':
                break
            else:
                nameList.append(currentName)
                i +=1
        nameList.sort()
        for i in range(len(nameList)):
            listBoxVariable.insert(tkinter.END, nameList[i])


# def get_form():
#     return optionVar.trace_add('write', lambda *args: print(optionVar.get()))


# def searchItem():
#     listBoxVariable.delete(0, 'end')
#     display = searchEntry.get() + ' not found.'
#     for j in range(len(nameList)):
#         currentItem = nameList[j]
#         if currentItem == searchEntry.get():
#             display = currentItem
#     listBoxVariable.insert(tkinter.END, display)

root.mainloop()
