#! python3


# TODO: figure out how it's possible that 'programToStart' gets called more than once
# read values and assign variables to the names and ranks


import tkinter
from tkinter import OptionMenu, Tk, Variable, mainloop, TOP, ttk
from tkinter.ttk import Button
from typing import ValuesView
import openpyxl
import os
import tkinter as tk
from ctypes import pythonapi
from sys import setswitchinterval
import pyautogui, time
from tkinter import ttk
import datetime
from datetime import datetime
from tkinter import *
pyautogui.FAILSAFE = True

delayTime = 3
interval_time = 0.0
path = r"C:\Users\skype\OneDrive\Documents\fillOutPDF_folder\ALPHA_ROSTER_FIELDS.xlsm"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
NORM_FONT = ("Helvetica", 10)


def create_GUI():
    # root = tkinter.Tk()
    root = Tk()
    # root.geometry('450x250')
    root["height"] = 250
    root["width"] = 450
    # root.geometry('450x250')
    root.title("Automated Form Filler")
    root.grid_propagate(False)
    root.config(bg='#FF2400')
    # root.resizable(False, False) 
    return root
root = create_GUI()
# def main():
#     w1 = Tk()
#     w1["height"] = 400;
#     w1["width"] = 500;
#     w1.title("Gui")
#     f1 = Frame(w1)
#     f1.pack_propagate(False)
#     f1["height"] = w1["height"];
#     f1["width"] = w1["width"];
#     f1.pack()
#     p1 = Button(f1)
#     p1["borderwidth"] = 6
#     p1["text"] = "esci"
#     p1["background"] = "red"
#     p1["command"] = f1.quit
#     p1.pack()
#     w1.mainloop()
# main()
def create_left_frame(root):
    leftSideFrameVariable = tkinter.Frame(root)
    leftSideFrameVariable.grid(row = 0, column = 0)
    # leftSideFrameVariable["height"] = 250
    # leftSideFrameVariable["width"] = 450
    # leftSideFrameVariable.grid_propagate(False)
    # leftSideFrameVariable.pack()
    return leftSideFrameVariable
leftSideFrameVariable = create_left_frame(root)

def create_right_frame(root):
    rightSideFrameVariable = tkinter.Frame(root)
    rightSideFrameVariable.grid(row = 0, column = 1)
    # leftSideFrameVariable["height"] = 250
    # leftSideFrameVariable["width"] = 450
    # rightSideFrameVariable.pack_propagate(False)
    # leftSideFrameVariable.pack()
    scrollbarVariable = tkinter.Scrollbar(
    rightSideFrameVariable,
    orient = tkinter.VERTICAL)
    listBoxVariable = tkinter.Listbox(
    rightSideFrameVariable,
    width = 25,
    yscrollcommand = scrollbarVariable.set,
    selectmode = tkinter.EXTENDED)
    scrollbarVariable.config(command = listBoxVariable.yview)
    scrollbarVariable.pack(side = tkinter.RIGHT, fill = tkinter.Y)
    listBoxVariable.pack()
    return listBoxVariable
# create_right_frame(root)
listBoxVariable = create_right_frame(root)

buttonVariable = tkinter.IntVar()
optionVar = tkinter.StringVar()
forms = ('Form 910 (AB - TSgt EPR)', 'Form 911 (MSgt - SMSgt EPR)', 'Form 4 (Reenlistment)')
ranks = ['AB', 'A1C', 'SrA', 'SSgt', 'TSgt', 'MSgt', 'SMSgt']
nameList = []
rankList = []


def determine_program_to_run():
    # pass
    optionMenuValue = optionVar.get()
    if optionMenuValue == 'Form 910 (AB - TSgt EPR)':
        programToStart = 'AF_form_910'
    elif optionMenuValue == 'Form 911 (MSgt - SMSgt EPR)':
        programToStart = 'AF_form_911'
    elif optionMenuValue == 'Form 4 (Reenlistment)':
        programToStart = 'Form_4_reenlistment'
    else:
        # error_message_popup("You must select a form")
        # textDisplayed = 'You must select a form'
        # processingLabel.configure(text = textDisplayed)
        programToStart = ''
    # print(programToStart)
    return programToStart

formOptionMenu = ttk.OptionMenu(
    leftSideFrameVariable, 
    optionVar, 
    'Choose form', 
    *forms)
    # ,
    # command = print_value_of_dropdown_menu) 
buttonLabel = tkinter.Label(
    leftSideFrameVariable,
    text = 'Choose:')
button1 = tkinter.Radiobutton(
    leftSideFrameVariable,
    text = 'By Rank',
    variable = buttonVariable,
    value = 1,
    justify='left',
    command=lambda: put_either_rank_or_names_into_listbox(button1))
button2 = tkinter.Radiobutton(
    leftSideFrameVariable,
    text = 'By Name',
    variable = buttonVariable,
    value = 2,
    justify='left',
    command=lambda: put_either_rank_or_names_into_listbox(button2))
runButton = tkinter.Button(
    leftSideFrameVariable,
    text = 'Run Program',
    command = lambda: runProgram(processingLabel))
processingLabel = tkinter.Label(
        leftSideFrameVariable,
        text = '')

def left_side_button_placement(formOptionMenu, button1, button2, runButton, processingLabel):
    formOptionMenu.grid(row = 0, column = 0, padx = 40, pady = 10)
    button1.grid(row = 3, column = 0)
    button2.grid(row = 4, column = 0)
    runButton.grid(row = 5, column = 0, padx = 100, pady = 10)
    processingLabel.grid(row = 6, column = 0)
left_side_button_placement(formOptionMenu, button1, button2, runButton, processingLabel)



# def runProgram(processingLabel): 
#     programToStart = determine_program_to_run()
#     print(programToStart) 
#     formToCall = "Write_" + programToStart #need to change names
#     rankOrName = choose_radio_button()
#     print(rankOrName)
#     selectedItems = []
#     for i in listBoxVariable.curselection():
#         selectedItem = listBoxVariable.get(i)
#         selectedItems.append(selectedItem)

#     if programToStart == '':
#         textDisplayed = 'You must select a form'
#     else:
#         if buttonVariable.get() != 1 or 2:
#             textDisplayed = 'You must make at least 1 choice of rank or name'
#         else:
#             if selectedItems == []:
#                 textDisplayed = ("Make at least 1 %s choice" % (rankOrName))
#     print(selectedItems)
#     processingLabel.configure(text = textDisplayed)
#     if programToStart != '' and rankOrName != '' and selectedItems != []:
#         if programToStart == 'AF_form_910':
#             Write_AF_form_910()
#     return rankOrName, nameList, rankList
# This method places either the ranks or the names into the field once either radio button is pressed.  Called as an argument by the radio buttons
def put_either_rank_or_names_into_listbox(buttonPressed):
    listBoxVariable.delete(0, tkinter.END)

    if buttonPressed == button1:
        
        for j in range(len(ranks)):
            listBoxVariable.insert(tkinter.END, ranks[j])
    
    elif buttonPressed == button2:
        names = []      #this needs to be its own method
        i = 2
        while str((sheet_obj.cell(row = i, column = 1)).value) != '':
            currentName = str((sheet_obj.cell(row = i, column = 1)).value)
            if currentName == 'None':
                break
            else:
                names.append(currentName)
                i +=1
        names.sort()
        for i in range(len(names)):
            listBoxVariable.insert(tkinter.END, names[i])

def choose_radio_button():
    if buttonVariable.get() == 1:
        rankOrName = 'rank'
        textDisplayed = rankOrName
        # print(rankOrName)
    elif buttonVariable.get() == 2:
        rankOrName = 'name'   
        textDisplayed = rankOrName 
        # print(rankOrName)   
    else:
        textDisplayed = 'You must choose either by rank or by name'
        rankOrName = ''
    processingLabel.configure(text = textDisplayed)
    # processingLabel.configure(text = 'You must choose either by rank or by name')
    return rankOrName
    #    error_message_popup("You must select either by name or by rank") 

def error_message_popup(msg):
    popup = tk.Tk()
    popup.wm_title("!")
    label = ttk.Label(popup, text=msg, font=NORM_FONT)
    label.pack(side="top", fill="x", pady=100, padx=200)
    B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
    B1.pack()
    popup.mainloop()

# root.mainloop()


##########################################      Form_910        ############################################




# NORM_FONT = ("Helvetica", 10)
def Write_AF_form_910(rankOrName, rankList, nameList):
    pdfName = 'main_af-form-910-enlisted-performance-report-ab-thru-tsgt.pdf'
   
    pdfCommand = 'start ' + pdfName
    

#######TODO#############
#determine if it's rank or name
#make a for loop to go through the array
# make another for loop for the entire program with the length of the array
# search for the appropriate characteristics to create the forms
# thingToSearch = rankOrName + "List"

    # if rankOrName == 'name': 
    #     attributeList = nameList
    # elif rankOrName == 'rank':
    #     attributeList = rankList
    #     # for n in nameList:



    i = 2
    while (sheet_obj.cell(row = i, column = 1)).value != None:
        
        name = (sheet_obj.cell(row = i, column = 1)).value
        print(name)
        for names in nameList:  ###############################################this for loop isn't being called for some reason
            if name == names:
                os.system(pdfCommand)
                ssn = sheet_obj.cell(row = i, column = 2)
                grade = (sheet_obj.cell(row = i, column = 3)).value

                ranks = {
                    'AB': 1,
                    'AMN': 2,
                    'A1C': 3,
                    'SRA': 4,
                    'SGT': 6,
                    'TSG': 8
                }
            
                # get() method of dictionary data type returns
                # value of passed argument if it is present
                # in dictionary otherwise second argument will
                # be assigned as default value of passed argument
                grade_key_presses = ranks.get(grade, 0)
                dafsc = (sheet_obj.cell(row = i, column = 11)).value
                command = "24 Special Tactics Squadron"
                pas = (sheet_obj.cell(row = i, column = 5)).value
                srid = '9999'
                report_start = ((sheet_obj.cell(row = i, column = 38)).value).date()
                report_end = ((sheet_obj.cell(row = i, column = 32)).value).date()
                supervisor = sheet_obj.cell(row=i, column=28)
                supervisor = supervisor.value
                ssn = ssn.value   #gets value from the ssn field
                ssn = ssn.replace("-", "")  #removes dashes from the ssn   #<class 'openpyxl.cell.cell.Cell'>
                date1 = report_start
                report_start = str(report_start)   #<class 'str'>
                new_report_start = datetime.strptime(report_start,'%Y-%m-%d').strftime('%d-%b-%Y')
                date2 = report_end
                report_end = str(report_end)
                new_report_end = datetime.strptime(report_end,'%Y-%m-%d').strftime('%d-%b-%Y')
                
                days_non_rated = 0
                days_supervised = (date2 - date1).days
                duty_title = (sheet_obj.cell(row = i, column = 8)).value
                
                #find supervisor information:
                # for j in range(2, 4):
                j = 2
                while (sheet_obj.cell(row = i, column = 1)).value != "":
                    print((sheet_obj.cell(row = i, column = 1)).value)
                    supervisor_name = ((sheet_obj.cell(row=j, column=1)).value)
                    shortened_supervisor_name = supervisor_name.replace(',', '').rsplit(' ', 1)[0]
                    if shortened_supervisor_name != supervisor:
                        pass
                    else:
                        supervisor_rank = (sheet_obj.cell(row=j, column=3)).value
                        supervisor_ssn = (sheet_obj.cell(row=j, column=2)).value
                        supervisor_last_four = repr(supervisor_ssn)[-5:-1]    #the last position is a quote mark because supervisor_last_four is a string
                        supervisor_information = ("%s, %s, USAF\n24 Special Tactics Squadron, AFSOC, Pope AAF, NC" % (supervisor_name, supervisor_rank))
                        supervisor_duty_title = (sheet_obj.cell(row = j, column = 8).value)
                        break  
                    j += 1
                time.sleep(delayTime*3)


                #### write everything to the PDF ######
                pyautogui.write(f'%s\t' % (name), interval=interval_time)
                pyautogui.write('%s\t' % (ssn), interval=interval_time)
                pyautogui.press('down', presses=grade_key_presses, interval=interval_time)
                pyautogui.press('tab', interval=interval_time)
                pyautogui.write('%s\t' % (dafsc), interval=interval_time)
                pyautogui.write('%s\t' % (command), interval=interval_time)
                pyautogui.write('%s\t' % (pas), interval=interval_time)
                pyautogui.write('%s\t' % (srid), interval=interval_time)
                pyautogui.write('%s\t' % (new_report_start), interval=interval_time)
                pyautogui.write('%s\t' % (new_report_end), interval=interval_time)
                pyautogui.sleep(2)
                pyautogui.press('enter', interval=interval_time)
                pyautogui.write('%s\t' % (days_non_rated), interval=interval_time)
                pyautogui.write('%s\t' % (days_supervised), interval=interval_time)

                pyautogui.press('down', interval=interval_time)       #sets reason for report to 'annual'.  
                pyautogui.press('tab', interval=interval_time)
                pyautogui.write('%s' % (duty_title), interval=interval_time)
                pyautogui.press('tab', presses=4, interval=interval_time)
                pyautogui.press('space, interval=interval_time')         
                pyautogui.press('tab', presses=6, interval=interval_time)
                pyautogui.press('space', interval=interval_time)         
                pyautogui.press('tab', presses=6, interval=interval_time)
                pyautogui.press('space', interval=interval_time) 
                pyautogui.press('tab', presses=6, interval=interval_time)
                pyautogui.press('space', interval=interval_time) 
                pyautogui.press('tab', presses=3, interval=interval_time)

                pyautogui.write('%s\t' % (supervisor_information), interval=interval_time)
                pyautogui.write('%s\t\t' % (supervisor_duty_title), interval=interval_time)
                pyautogui.write('%s' % (supervisor_last_four), interval=interval_time)
            


                #save the document
                pyautogui.hotkey('ctrl', 's')
                pyautogui.sleep(delayTime)
                pyautogui.press('enter')
                pyautogui.sleep(delayTime)
                pyautogui.write(name + ' EPR')
                pyautogui.press('enter')
                pyautogui.hotkey('ctrl', 'f4')

        i += 1


def runProgram(processingLabel): 
    programToStart = determine_program_to_run()
    # print(programToStart) 
    formToCall = "Write_" + programToStart #need to change names
    rankOrName = choose_radio_button()
    # print(rankOrName)
    selectedItems = []
    for i in listBoxVariable.curselection():
        selectedItem = listBoxVariable.get(i)
        selectedItems.append(selectedItem)

    if programToStart == '':
        textDisplayed = 'You must select a form'
    else:
        if buttonVariable.get() != 1 or 2:
            textDisplayed = 'You must make at least 1 choice of rank or name'
        else:
            if selectedItems == []:
                textDisplayed = ("Make at least 1 %s choice" % (rankOrName))
    # print(selectedItems)
    processingLabel.configure(text = textDisplayed)
    if programToStart != '' and rankOrName != '' and selectedItems != []:
        if programToStart == 'AF_form_910':
            Write_AF_form_910(rankOrName, rankList, nameList)
    return rankOrName, nameList, rankList

# def OperationCompleteMessage(msg):
#     popup = tk.Tk()
#     popup.wm_title("!")
#     label = ttk.Label(popup, text=msg, font=NORM_FONT)
#     label.pack(side="top", fill="x", pady=100, padx=500)
#     B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
#     B1.pack()
#     popup.mainloop()

# OperationCompleteMessage("Operation Complete!")


root.mainloop()

