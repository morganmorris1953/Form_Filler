from ctypes import pythonapi
import pyautogui, time
import os
import tkinter as tk
from tkinter import ttk
import csv
import openpyxl
import datetime
from datetime import datetime
from datetime import date


NORM_FONT = ("Helvetica", 10)
delayTime = 8

pdfName = 'main_af-form-911-enlisted-performance-report-msgt-thru-smsgt.pdf'
pdfCommand = 'start ' + pdfName
supervisor_information = 'test'
supervisor_ssn = '000000000'

path = r"C:\Users\skype\OneDrive\Documents\fillOutPDF_folder\ALPHA_ROSTER_FIELDS.xlsm"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active


for i in range(2, 4):
    

    os.system(pdfCommand)
    name = sheet_obj.cell(row = i, column = 1)
    ssn = sheet_obj.cell(row = i, column = 2)
    grade = sheet_obj.cell(row = i, column = 3)
    dafsc = sheet_obj.cell(row = i, column = 11)
    command = "24 Special Tactics Squadron"
    pas = sheet_obj.cell(row = i, column = 5)
    srid = '9999'
    report_start = sheet_obj.cell(row = i, column = 38)
    report_end = sheet_obj.cell(row = i, column = 32)
    supervisor = sheet_obj.cell(row=i, column=27)
    supervisor = str(supervisor)
    supervisor = supervisor.replace(',', '').rsplit(' ', 1)[0]
    
    #find supervisor information:
    for j in range(2, 4):
        if supervisor != (str(sheet_obj.cell(row=1, column=j))):
            print('sup not recognized')
            # pass
        else:
            supervisor_rank = sheet_obj.cell(row=j, column=3)
            supervisor_ssn = sheet_obj.cell(row=j, column=2)
            supervisor_ssn = supervisor_ssn.value   #gets value from the ssn field
            supervisor_last_four = int(repr(supervisor_ssn)[-4:])
            # supervisor_ssn = supervisor_ssn.replace("-", "")  #removes dashes from the ssn   #<class 'openpyxl.cell.cell.Cell'>
            supervisor_information = ("%s, %s, USAF, 24 Special Tactics Squadron, AFSOC, Pope AAF, NC" % (supervisor, supervisor_rank))
            print("Supervisor information is received")
            print(supervisor_ssn, supervisor_rank, supervisor_information)
            break

            #name, rank, service, org, command, location
    
    ssn = ssn.value   #gets value from the ssn field
    ssn = ssn.replace("-", "")  #removes dashes from the ssn   #<class 'openpyxl.cell.cell.Cell'>
    report_start = report_start.value   #<class 'datetime.datetime'>
    report_start = report_start.date()   #<class 'datetime.date'>
    date1 = report_start
    report_start = str(report_start)   #<class 'str'>
    new_report_start = datetime.strptime(report_start,'%Y-%m-%d').strftime('%d-%b-%Y')
    report_end = report_end.value
    report_end = report_end.date()
    date2 = report_end
    report_end = str(report_end)
    new_report_end = datetime.strptime(report_end,'%Y-%m-%d').strftime('%d-%b-%Y')
    days_non_rated = 0
    days_supervised = (date2 - date1).days

    time.sleep(delayTime)


    #### write everything to the PDF ######
    pyautogui.write(f'%s\t' % (name.value))
    pyautogui.write('%s\t' % (ssn))
    pyautogui.write('%s\t' % (grade.value))
    pyautogui.write('%s\t' % (dafsc.value))
    pyautogui.write('%s\t' % (command))
    pyautogui.write('%s\t' % (pas.value))
    pyautogui.write('%s\t' % (srid))
    pyautogui.write('%s\t' % (new_report_start))
    pyautogui.write('%s\t' % (new_report_end))
    pyautogui.write('%s\t' % (days_non_rated))
    pyautogui.write('%s\t' % (days_supervised))

    pyautogui.press('down')       #sets reason for report to 'annual'.  
    pyautogui.press('tab', presses=5)
    pyautogui.press('space')         
    pyautogui.press('tab', presses=6)
    pyautogui.press('space')         
    pyautogui.press('tab', presses=7)
    pyautogui.press('space') 
    pyautogui.press('tab', presses=3)

    pyautogui.write('%s\t\t' % (supervisor_information))
    # pyautogui.write('%s\t' % (date_today))        #date_today is causing issues
    pyautogui.write('%s' % (supervisor_ssn))
   


    #save the document
    pyautogui.hotkey('ctrl', 's')
    pyautogui.sleep(4)
    pyautogui.press('enter')
    pyautogui.sleep(2)
    pyautogui.write(name.value + ' EPR')
    pyautogui.press('enter')
    pyautogui.hotkey('altleft', 'f4')

    # print(name.value, ssn, grade.value, dafsc.value, command, pas.value, srid, new_report_start, new_report_end)




# def OperationCompleteMessage(msg):
#     popup = tk.Tk()
#     popup.wm_title("!")
#     label = ttk.Label(popup, text=msg, font=NORM_FONT)
#     label.pack(side="top", fill="x", pady=100, padx=500)
#     B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
#     B1.pack()
#     popup.mainloop()

# OperationCompleteMessage("Operation Complete!")

