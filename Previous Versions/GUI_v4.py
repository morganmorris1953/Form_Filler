#! python3


# TODO: figure out how it's possible that 'programToStart' gets called more than once
# read values and assign variables to the names and ranks


import tkinter
from tkinter import OptionMenu, Tk, Variable, mainloop, TOP, ttk
from tkinter.ttk import Button
from typing import ValuesView
import openpyxl
import os
import tkinter as tk
# import AF_form_910_v3     this automatically runs the program.  weird.
# import AF_form_911_v2


path = r"C:\Users\skype\OneDrive\Documents\fillOutPDF_folder\ALPHA_ROSTER_FIELDS.xlsm"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
NORM_FONT = ("Helvetica", 10)


def create_GUI():
    root = tkinter.Tk()
    root.geometry('450x250')
    root.title("Automated Form Filler")
    root.config(bg='#F2B90C')
    return root
root = create_GUI()

def create_left_frame(root):
    leftSideFrameVariable = tkinter.Frame(root)
    leftSideFrameVariable.grid(row = 0, column = 0)
    # leftSideFrameVariable.config(bg='#F2B90C')
    return leftSideFrameVariable
leftSideFrameVariable = create_left_frame(root)

def create_right_frame(root):
    rightSideFrameVariable = tkinter.Frame(root)
    rightSideFrameVariable.grid(row = 0, column = 1)
    scrollbarVariable = tkinter.Scrollbar(
    rightSideFrameVariable,
    orient = tkinter.VERTICAL)
    listBoxVariable = tkinter.Listbox(
    rightSideFrameVariable,
    width = 25,
    yscrollcommand = scrollbarVariable.set,
    selectmode = tkinter.EXTENDED)
    scrollbarVariable.config(command = listBoxVariable.yview)
    scrollbarVariable.pack(side = tkinter.RIGHT, fill = tkinter.Y)
    listBoxVariable.pack()
    return listBoxVariable
# create_right_frame(root)
listBoxVariable = create_right_frame(root)

buttonVariable = tkinter.IntVar()
optionVar = tkinter.StringVar()
forms = ('Form 910 (AB - TSgt EPR)', 'Form 911 (MSgt - SMSgt EPR)', 'Form 4 (Reenlistment)')
ranks = ['AB', 'A1C', 'SrA', 'SSgt', 'TSgt', 'MSgt', 'SMSgt']
nameList = []
rankList = []


def determine_program_to_run():
    # pass
    optionMenuValue = optionVar.get()
    if optionMenuValue == 'Form 910 (AB - TSgt EPR)':
        programToStart = 'AF_form_910_v3.py'
    elif optionMenuValue == 'Form 911 (MSgt - SMSgt EPR)':
        programToStart = 'AF_form_911_v2.py'
    elif optionMenuValue == 'Form 4 (Reenlistment)':
        programToStart = 'Form_4_reenlistment_v1.py'
    else:
        # error_message_popup("You must select a form")
        # textDisplayed = 'You must select a form'
        # processingLabel.configure(text = textDisplayed)
        programToStart = ''
    # print(programToStart)
    return programToStart

formOptionMenu = ttk.OptionMenu(
    leftSideFrameVariable, 
    optionVar, 
    'Choose form', 
    *forms)
    # ,
    # command = print_value_of_dropdown_menu) 
buttonLabel = tkinter.Label(
    leftSideFrameVariable,
    text = 'Choose:')
button1 = tkinter.Radiobutton(
    leftSideFrameVariable,
    text = 'By Rank',
    variable = buttonVariable,
    value = 1,
    justify='left',
    command=lambda: put_either_rank_or_names_into_listbox(button1))
button2 = tkinter.Radiobutton(
    leftSideFrameVariable,
    text = 'By Name',
    variable = buttonVariable,
    value = 2,
    justify='left',
    command=lambda: put_either_rank_or_names_into_listbox(button2))
runButton = tkinter.Button(
    leftSideFrameVariable,
    text = 'Run Program',
    command = lambda: runProgram(processingLabel))
processingLabel = tkinter.Label(
        leftSideFrameVariable,
        text = '')

def left_side_button_placement(formOptionMenu, button1, button2, runButton, processingLabel):
    formOptionMenu.grid(row = 0, column = 0, padx = 40, pady = 10)
    button1.grid(row = 3, column = 0)
    button2.grid(row = 4, column = 0)
    runButton.grid(row = 5, column = 0, padx = 100, pady = 10)
    processingLabel.grid(row = 6, column = 0)
left_side_button_placement(formOptionMenu, button1, button2, runButton, processingLabel)



def runProgram(processingLabel): 
    programToStart = determine_program_to_run()
    print(programToStart) 
    rankOrName = choose_radio_button()
    print(rankOrName)
    selectedItems = []
    for i in listBoxVariable.curselection():
        selectedItem = listBoxVariable.get(i)
        selectedItems.append(selectedItem)

    if programToStart == '':
        textDisplayed = 'You must select a form'
    else:
        if buttonVariable.get() != 1 or 2:
            textDisplayed = 'You must make at least 1 choice of rank or name'
        else:
            # selectedItems = []
            # for i in listBoxVariable.curselection():
            #     selectedItem = listBoxVariable.get(i)
            #     selectedItems.append(selectedItem)
            if selectedItems == []:
                textDisplayed = ("Make at least 1 %s choice" % (rankOrName))
            # else:
    print(selectedItems)
    processingLabel.configure(text = textDisplayed)

    # selectedItems = []
    # for i in listBoxVariable.curselection():
    #     selectedItem = listBoxVariable.get(i)
    #     selectedItems.append(selectedItem)
    # if selectedItems == []:
    #     # print("Make at least 1 choice")
    #     if buttonVariable.get() == 1 or 2:
    #         textDisplayed = 'You must make at least 1 choice of rank or name'
    #         processingLabel.configure(text = textDisplayed)
    # else:
    #     print(selectedItems)



# This method places either the ranks or the names into the field once either radio button is pressed.  Called as an argument by the radio buttons
def put_either_rank_or_names_into_listbox(buttonPressed):
    listBoxVariable.delete(0, tkinter.END)

    if buttonPressed == button1:
        
        for j in range(len(ranks)):
            listBoxVariable.insert(tkinter.END, ranks[j])
    
    elif buttonPressed == button2:
        names = []      #this needs to be its own method
        i = 2
        while str((sheet_obj.cell(row = i, column = 1)).value) != '':
            currentName = str((sheet_obj.cell(row = i, column = 1)).value)
            if currentName == 'None':
                break
            else:
                names.append(currentName)
                i +=1
        names.sort()
        for i in range(len(names)):
            listBoxVariable.insert(tkinter.END, names[i])

def choose_radio_button():
    if buttonVariable.get() == 1:
        rankOrName = 'rank'
        textDisplayed = rankOrName
        # print(rankOrName)
    elif buttonVariable.get() == 2:
        rankOrName = 'name'   
        textDisplayed = rankOrName 
        # print(rankOrName)   
    else:
        textDisplayed = 'You must choose either by rank or by name'
        rankOrName = ''
    processingLabel.configure(text = textDisplayed)
    # processingLabel.configure(text = 'You must choose either by rank or by name')
    return rankOrName
    #    error_message_popup("You must select either by name or by rank") 

def error_message_popup(msg):
    popup = tk.Tk()
    popup.wm_title("!")
    label = ttk.Label(popup, text=msg, font=NORM_FONT)
    label.pack(side="top", fill="x", pady=100, padx=200)
    B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
    B1.pack()
    popup.mainloop()

root.mainloop()
