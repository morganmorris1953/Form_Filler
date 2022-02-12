from ctypes import pythonapi
import pyautogui, time
import os
import tkinter as tk
from tkinter import ttk
import csv
import openpyxl
import datetime
from datetime import datetime
from datetime import date


NORM_FONT = ("Helvetica", 10)
delayTime = 8

pdfName = 'main_af-form-911-enlisted-performance-report-msgt-thru-smsgt.pdf'
pdfCommand = 'start ' + pdfName


path = r"C:\Users\skype\OneDrive\Documents\fillOutPDF_folder\ALPHA_ROSTER_FIELDS.xlsm"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active


for i in range(2, 4):
    

    os.system(pdfCommand)
    name = (sheet_obj.cell(row = i, column = 1)).value
    ssn = sheet_obj.cell(row = i, column = 2)
    grade = (sheet_obj.cell(row = i, column = 3)).value
    if grade == 'MSG':
        grade_key_presses = 2
    else:
        grade_key_presses = 4
    dafsc = (sheet_obj.cell(row = i, column = 11)).value
    command = "24 Special Tactics Squadron"
    pas = (sheet_obj.cell(row = i, column = 5)).value
    srid = '9999'
    report_start = ((sheet_obj.cell(row = i, column = 38)).value).date()
    report_end = ((sheet_obj.cell(row = i, column = 32)).value).date()
    supervisor = sheet_obj.cell(row=i, column=28)
    supervisor = supervisor.value
    ssn = ssn.value   #gets value from the ssn field
    ssn = ssn.replace("-", "")  #removes dashes from the ssn   #<class 'openpyxl.cell.cell.Cell'>
    # report_start = report_start.value   #<class 'datetime.datetime'>
    # report_start = report_start.date()   #<class 'datetime.date'>
    date1 = report_start
    report_start = str(report_start)   #<class 'str'>
    new_report_start = datetime.strptime(report_start,'%Y-%m-%d').strftime('%d-%b-%Y')
    # report_end = report_end.value
    # report_end = report_end.date()
    date2 = report_end
    report_end = str(report_end)
    new_report_end = datetime.strptime(report_end,'%Y-%m-%d').strftime('%d-%b-%Y')
    
    days_non_rated = 0
    days_supervised = (date2 - date1).days
    duty_title = (sheet_obj.cell(row = i, column = 8)).value
    
   
    
    #find supervisor information:
    for j in range(2, 4):
        supervisor_name = ((sheet_obj.cell(row=j, column=1)).value)
        shortened_supervisor_name = supervisor_name.replace(',', '').rsplit(' ', 1)[0]
        if shortened_supervisor_name != supervisor:
            pass
        else:
            supervisor_rank = (sheet_obj.cell(row=j, column=3)).value
            supervisor_ssn = (sheet_obj.cell(row=j, column=2)).value
            supervisor_last_four = repr(supervisor_ssn)[-5:-1]    #the last position is a quote mark because supervisor_last_four is a string
            supervisor_information = ("%s, %s, USAF\n24 Special Tactics Squadron, AFSOC, Pope AAF, NC" % (supervisor_name, supervisor_rank))
            supervisor_duty_title = (sheet_obj.cell(row = j, column = 8).value)
            break  

    time.sleep(delayTime)


    #### write everything to the PDF ######
    pyautogui.write(f'%s\t' % (name))
    pyautogui.write('%s\t' % (ssn))
    # pyautogui.write('%s\t' % (grade))
    pyautogui.press('down', presses=grade_key_presses)
    pyautogui.press('tab')
    pyautogui.write('%s\t' % (dafsc))
    pyautogui.write('%s\t' % (command))
    pyautogui.write('%s\t' % (pas))
    pyautogui.write('%s\t' % (srid))
    pyautogui.write('%s\t' % (new_report_start))
    pyautogui.write('%s\t' % (new_report_end))
    pyautogui.sleep(2)
    pyautogui.press('enter')
    pyautogui.write('%s\t' % (days_non_rated))
    pyautogui.write('%s\t' % (days_supervised))

    pyautogui.press('down')       #sets reason for report to 'annual'.  
    pyautogui.press('tab')
    pyautogui.write('%s' % (duty_title))
    pyautogui.press('tab', presses=4)
    pyautogui.press('space')         
    pyautogui.press('tab', presses=6)
    pyautogui.press('space')         
    pyautogui.press('tab', presses=7)
    pyautogui.press('space') 
    pyautogui.press('tab', presses=3)

    pyautogui.write('%s\t' % (supervisor_information))
    pyautogui.write('%s\t\t' % (supervisor_duty_title))
    pyautogui.write('%s' % (supervisor_last_four))
   


    #save the document
    pyautogui.hotkey('ctrl', 's')
    pyautogui.sleep(4)
    pyautogui.press('enter')
    pyautogui.sleep(2)
    pyautogui.write(name + ' EPR')
    pyautogui.press('enter')
    pyautogui.hotkey('ctrl', 'f4')


    # print(name.value, ssn, grade.value, dafsc.value, command, pas.value, srid, new_report_start, new_report_end)




# def OperationCompleteMessage(msg):
#     popup = tk.Tk()
#     popup.wm_title("!")
#     label = ttk.Label(popup, text=msg, font=NORM_FONT)
#     label.pack(side="top", fill="x", pady=100, padx=500)
#     B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
#     B1.pack()
#     popup.mainloop()

# OperationCompleteMessage("Operation Complete!")

